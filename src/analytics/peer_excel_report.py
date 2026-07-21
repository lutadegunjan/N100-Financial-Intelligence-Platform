"""
N100 Financial Intelligence Platform
Sprint 3 - Day 20
Peer Comparison Excel Report Generator
"""

import sqlite3
import pandas as pd
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


DB_PATH = "db/nifty100.db"

OUTPUT_FILE = "output/peer_comparison.xlsx"


METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover"
]


def load_peer_data():

    conn = sqlite3.connect(DB_PATH)


    percentile = pd.read_sql(
        """
        SELECT *
        FROM peer_percentiles
        """,
        conn
    )


    companies = pd.read_sql(
        """
        SELECT id,
               company_name
        FROM companies
        """,
        conn
    )


    conn.close()


    df = percentile.merge(
        companies,
        left_on="company_id",
        right_on="id",
        how="left"
    )


    return df



def create_report():


    os.makedirs(
        "output",
        exist_ok=True
    )


    df = load_peer_data()


    writer = pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    )


    peer_groups = (
        df["peer_group_name"]
        .dropna()
        .unique()
    )


    for group in peer_groups:


        group_df = df[
            df["peer_group_name"]
            ==
            group
        ]


        pivot_value = group_df.pivot_table(
            index=[
                "company_id",
                "company_name"
            ],
            columns="metric",
            values="value"
        )


        pivot_rank = group_df.pivot_table(
            index=[
                "company_id",
                "company_name"
            ],
            columns="metric",
            values="percentile_rank"
        )


        pivot_rank.columns = [
            f"{col}_percentile"
            for col in pivot_rank.columns
        ]


        final = pd.concat(
            [
                pivot_value,
                pivot_rank
            ],
            axis=1
        )


        final = final.reset_index()


        median = pd.DataFrame(
            [
                final.select_dtypes(
                    include="number"
                )
                .median()
            ]
        )


        median.insert(
            0,
            "company_id",
            "MEDIAN"
        )


        median.insert(
            1,
            "company_name",
            "Peer Group Median"
        )


        final = pd.concat(
            [
                final,
                median
            ],
            ignore_index=True
        )


        sheet_name = (
            group[:31]
        )


        final.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False
        )


    writer.close()


    format_excel()


    print(
        "Created:",
        OUTPUT_FILE
    )



def format_excel():


    wb = load_workbook(
        OUTPUT_FILE
    )


    green = PatternFill(
        "solid",
        fgColor="C6EFCE"
    )

    yellow = PatternFill(
        "solid",
        fgColor="FFEB9C"
    )

    red = PatternFill(
        "solid",
        fgColor="FFC7CE"
    )


    gold = PatternFill(
        "solid",
        fgColor="FFD966"
    )


    for ws in wb:


        for row in ws.iter_rows():

            for cell in row:


                if (
                    isinstance(
                        cell.value,
                        float
                    )
                    and
                    "percentile"
                    in ws.cell(
                        1,
                        cell.column
                    ).value
                ):


                    if cell.value >= 75:

                        cell.fill = green

                    elif cell.value <= 25:

                        cell.fill = red

                    else:

                        cell.fill = yellow



        for row in ws.iter_rows():

            if (
                row[0].value
                ==
                "MEDIAN"
            ):

                for cell in row:

                    cell.font = Font(
                        bold=True
                    )


    wb.save(
        OUTPUT_FILE
    )



if __name__ == "__main__":

    create_report()